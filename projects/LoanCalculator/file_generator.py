import openpyxl

from logic import *


def excel_report_generator():
    """
    生成Excel报告
    """
    P, n, i, Y, S, I, Yp, Yi, Pr = equal_principal_and_interest(P=285000, i=3.1, n=276)
    workbook = openpyxl.load_workbook('报告模板.xlsx')
    sheet = workbook['基本信息']
    sheet['C3'] = P
    sheet['C4'] = n
    sheet['C5'] = i
    sheet['C6'] = '等额本息'
    sheet['C7'] = Y
    sheet['C8'] = 0
    sheet['C9'] = S
    sheet['C10'] = I

    sheet = workbook['原还款明细']
    for i in range(1, n + 1):
        sheet.cell(row=i + 1, column=1).value = f'第{i:03d}期'
        sheet.cell(row=i + 1, column=2).value = Yp[i - 1]
        sheet.cell(row=i + 1, column=3).value = Yi[i - 1]
        sheet.cell(row=i + 1, column=4).value = Y
        sheet.cell(row=i + 1, column=5).value = Pr[i - 1]

    workbook.save('报告.xlsx')
    pass


if __name__ == '__main__':
    excel_report_generator()
