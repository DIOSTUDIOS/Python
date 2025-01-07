from configparser import ConfigParser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from recorder import Recorder

import pymysql
import numbers


recorder = Recorder()


def get_database_config():
    config = ConfigParser()
    config.read('config.ini', encoding='utf-8')

    databaseInfo = {
        'host': config.get('database', 'host'),
        'port': config.get('database', 'port'),
        'database': config.get('database', 'database'),
        'username': config.get('database', 'username'),
        'password': config.get('database', 'password')
    }

    return databaseInfo


def get_sql(tableName):
    config = ConfigParser()
    config.read('config.ini', encoding='utf-8')

    sqls = {
        'ckmx': config.get('sqls', 'ckmx'),
        'cpmx': config.get('sqls', 'cpmx'),
        'ddmx': config.get('sqls', 'ddmx'),
        'htmx': config.get('sqls', 'htmx'),
        'zpsr': config.get('sqls', 'zpsr'),
        'yflr': config.get('sqls', 'yflr'),
        'fpsr': config.get('sqls', 'fpsr'),
        'lrmx': config.get('sqls', 'lrmx'),
        'lrhz': config.get('sqls', 'lrhz')
    }

    return sqls[tableName]


def get_datasheet(tableName='ckmx'):
    databaseInfo = get_database_config()

    connection = pymysql.connect(host=databaseInfo['host'],
                                 port=int(databaseInfo['port']),
                                 database=databaseInfo['database'],
                                 user=databaseInfo['username'],
                                 password=databaseInfo['password'])
    cursor = connection.cursor()
    recorder.record_log('info', '数据库连接开启!')
    recorder.record_log('info', '数据获取中...!')

    cursor.execute(get_sql(tableName))
    columnNames = [desc[0] for desc in cursor.description]
    dataSheet = cursor.fetchall()
    recorder.record_log('info', '数据获取成功!')

    cursor.close()
    connection.close()
    recorder.record_log('info', '数据库连接关闭!')

    return tuple(columnNames), dataSheet


def generate_excel(tableName='ckmx', documentName='excel'):
    wb = Workbook()
    sheet = wb.active

    columnNames, dataSheet = get_datasheet(tableName=tableName)
    recorder.record_log('info', '创建并写入Excel数据表...')

    for column in range(0, len(columnNames)):
        sheet.cell(row=1, column=column+1).value = columnNames[column]
        sheet.cell(row=1, column=column+1).font = Font(name='等线', bold=True,  size=10)
        sheet.cell(row=1, column=column+1).alignment = Alignment(horizontal='center', vertical='center')

    for row in range(0, len(dataSheet)):
        for column in range(0, len(dataSheet[row])):
            sheet.cell(row=row+2, column=column+1).value = dataSheet[row][column]
            sheet.cell(row=row+2, column=column+1).font = Font(name='等线', size=10)
            sheet.cell(row=row+2, column=column+1).alignment = Alignment(horizontal='center', vertical='center')

            if isinstance(sheet.cell(row=row+2, column=column+1).value, numbers.Number):
                sheet.cell(row=row+2, column=column+1).number_format = '0.00'
                sheet.cell(row=row+2, column=column+1).alignment = Alignment(horizontal='right', vertical='center')

                if sheet.cell(row=row+2, column=column+1).value < 0:
                    sheet.cell(row=row+2, column=column+1).font = Font(name='等线', size=10, color='ff0000')
                else:
                    pass
            else:
                pass

    sheet.freeze_panes = 'A2'

    wb.save(f'./{documentName}.xlsx')
    recorder.record_log('info', f'生成 Excel 文件 {documentName}.xlsx 成功!')


if __name__ == '__main__':
    generate_excel()
