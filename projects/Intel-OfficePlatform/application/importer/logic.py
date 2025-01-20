from datetime import datetime, timedelta

import os
import openpyxl
import sqlite3


# 导入通讯卡ICCID、开通时间、到期时间等信息到 simcard.db 数据库
def import_simcard_info(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    db = sqlite3.connect('simcard.db')
    cursor = db.cursor()
    try:
        for i in range(1, sheet.max_row+1):
            # 获取 iccid 卡号
            iccid = sheet.cell(row=i, column=1).value
            # 获取套餐开始时间
            start_time = sheet.cell(row=i, column=2).value
            # 获取套餐结束时间
            end_time = sheet.cell(row=i, column=3).value
            # 获取通讯卡运营商
            if str(iccid)[:6] in ['898600', '898602', '898604', '898607']:
                operator = '中国移动'
            elif str(iccid)[:6] in ['898601', '898606', '898609']:
                operator = '中国联通'
            elif str(iccid)[:6] in ['898603', '898611']:
                operator = '中国电信'
            else:
                operator = '其他'
            # 将获取到的信息插入数据库
            cursor.execute(r"INSERT INTO simcard(iccid,start_time,end_time,operator,is_used,deleted) VALUES('{0}','{1}','{2}','{3}','0','0')".format(iccid, start_time, end_time, operator))
            db.commit()
    except SystemExit:
        db.rollback()
    finally:
        cursor.close()
        db.close()

        return True

    pass


# 生成设备信息导入文件
def create_devices_file(file):
    # 打开选择的 excel 文件
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    # 设备编号列表
    deviceNumbers = []
    # 设备信息列表
    devices = []
    # 获取设备编号，并存入 deviceNumbers 列表
    for i in range(1, sheet.max_row + 1):
        deviceNumbers.append(sheet.cell(row=i, column=1).value)

    for deviceNumber in deviceNumbers:
        # 获取 设备类型
        if str(deviceNumber)[3] == '6':
            deviceType = '户阀'
        elif str(deviceNumber)[3] == '7':
            deviceType = '室温'
        elif str(deviceNumber)[3] == '8':
            deviceType = '管道测温'
        else:
            deviceType = '执行器'
        # 获取 供电方式
        if str(deviceNumber)[2] == '0':
            powerMode = '充电电池'
        elif str(deviceNumber)[2] == '1':
            powerMode = '长效电池'
        else:
            powerMode = '市电'
        # 获取 通讯类型
        if str(deviceNumber)[0:2] in ['20', '21']:
            if str(deviceNumber)[4:] > '49999':
                reportType = '电信NB'
            else:
                reportType = '移动NB'
        else:
            if str(deviceNumber)[4] == '0':
                reportType = '移动NB'
            elif str(deviceNumber)[4] == '1':
                reportType = '电信NB'
            elif str(deviceNumber)[4] == '2':
                reportType = '移动4G(Cat-1)'
            elif str(deviceNumber)[4] == '3':
                reportType = '电信4G(Cat-1)'
            elif str(deviceNumber)[4] == '4':
                reportType = '联通4G(Cat-1)'
            elif str(deviceNumber)[4] == '5':
                reportType = 'Modbus'
            else:
                reportType = 'LoRa'
        # 获取 输入电压
        if str(deviceNumber)[3] == '6':
            inputVol = '7.4VDC'
        elif str(deviceNumber)[3] == '7':
            inputVol = '220VAC'
        elif str(deviceNumber)[3] == '8':
            inputVol = '3.7VDC'
        else:
            inputVol = '7.4VDC'
        # 获取 规格型号
        if str(deviceNumber)[3] == '7':
            if str(deviceNumber)[0:2] not in ['20', '21']:
                if str(deviceNumber)[4] in ['0', '1']:
                    comType = 'N'
                else:
                    comType = 'C'
            else:
                comType = 'N'

            if str(deviceNumber)[0:2] not in ['20', '21']:
                if str(deviceNumber)[4] in ['0', '2']:
                    period = 'YD5'
                else:
                    period = 'DX5'
            else:
                if str(deviceNumber)[4:] > '49999':
                    period = 'DX5'
                else:
                    period = 'YD5'

            specModel = 'RTM-TX{0}F-86-{1}-{2}'.format(str(deviceNumber)[1], comType, period)
        elif str(deviceNumber)[3] == '8':
            if str(deviceNumber)[4] in ['0', '1']:
                comType = 'N'
            else:
                comType = 'C'

            if str(deviceNumber)[0:2] not in ['20', '21']:
                if str(deviceNumber)[4] in ['0', '2']:
                    period = 'YD5'
                else:
                    period = 'DX5'
            else:
                if str(deviceNumber)[4:] > '49999':
                    period = 'DX5'
                else:
                    period = 'YD5'

            specModel = 'PTM-TX{0}F-GD-{1}-{2}'.format(str(deviceNumber)[1], comType, period)
        elif str(deviceNumber)[3] == '9':
            if str(deviceNumber)[2] == '0':
                batType = 'DC2'
            elif str(deviceNumber)[2] == '1':
                batType = 'DC1'
            else:
                batType = 'AC'

            if str(deviceNumber)[4] in ['0', '1']:
                comType = 'N'
            else:
                comType = 'C'

            if str(deviceNumber)[0:2] not in ['20', '21']:
                if str(deviceNumber)[4] in ['0', '2']:
                    period = 'YD5'
                else:
                    period = 'DX5'
            else:
                if str(deviceNumber)[4:] > '49999':
                    period = 'DX5'
                else:
                    period = 'YD5'

            specModel = 'TXPF-ATT{0}F-{1}-{2}-{3}'.format(str(deviceNumber)[1], batType, comType, period)
        else:
            specModel = '其他'
        # 获取 批次号码
        if str(deviceNumber)[3] == '6':
            batchNumber = 'HF' + datetime.now().strftime('%y%m%d')
        elif str(deviceNumber)[3] == '7':
            batchNumber = 'SW' + datetime.now().strftime('%y%m%d')
        elif str(deviceNumber)[3] == '8':
            batchNumber = 'GD' + datetime.now().strftime('%y%m%d')
        else:
            batchNumber = 'ZX' + datetime.now().strftime('%y%m%d')
        # 生成格式化后的设备信息列表
        devices.append(['0', str(deviceNumber), deviceType, powerMode, reportType, inputVol, specModel, batchNumber])

    workbook = openpyxl.load_workbook(os.getcwd() + r'\设备模板.xlsx')
    sheet = workbook.active
    # 将每行设备信息追加到当前 sheet 中
    for device in devices:
        sheet.append(device)
    # 保存文件到当前位置
    workbook.save('设备导入文件.xlsx')

    return True


# 生成卡片信息导入文件
def create_simcard_file(file, year):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    # 通讯卡列表
    simcard = []

    db = sqlite3.connect('simcard.db')
    cursor = db.cursor()

    for i in range(1, sheet.max_row+1):
        # 获取通讯卡 iccid 号码
        iccid = ''.join(sheet.cell(row=i, column=2).value.split())
        # 获取设备编号
        device_no = sheet.cell(row=i, column=1).value
        # 获取通讯卡通讯方式
        if len(str(device_no)) == 9:
            deviceType = 'NB'
        else:
            if str(device_no)[4] in ['0', '1']:
                deviceType = 'NB'
            elif str(device_no)[4] in ['2', '3', '4']:
                deviceType = '4G'
            elif str(device_no)[4] == '5':
                deviceType = 'modbus'
            else:
                deviceType = 'LoRa'
        # 获取通讯卡运营商
        if str(iccid)[:6] in ['898600', '898602', '898604', '898607']:
            operator = '中国移动'
        elif str(iccid)[:6] in ['898601', '898606', '898609']:
            operator = '中国联通'
        elif str(iccid)[:6] in ['898603', '898611']:
            operator = '中国电信'
        else:
            operator = '其他'
        # 获取通讯卡套餐结束时间
        cursor.execute(r"SELECT end_time FROM simcard WHERE iccid='{0}'".format(iccid))
        result = cursor.fetchone()

        if result is None:
            limit = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            limit = result[0]
        # 更新通讯卡是否使用
        cursor.execute(r"UPDATE simcard SET is_used=1 WHERE iccid='{0}'".format(iccid))
        # 更新通讯卡绑定的设备编号
        cursor.execute(r"UPDATE simcard SET device_no='{0}' WHERE iccid='{1}'".format(device_no, iccid))
        # 通讯卡信息格式化
        simcard.append(['0', iccid, deviceType, operator, operator, device_no, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '正常', (datetime.now() + timedelta(days=(int(year)*365))).strftime('%Y-%m-%d %H:%M:%S'), limit])

    db.commit()

    cursor.close()
    db.close()

    workbook = openpyxl.load_workbook(os.getcwd() + '\卡片模板.xlsx')
    sheet = workbook.active

    for deviceNumber in simcard:
        sheet.append(deviceNumber)

    sheet.delete_rows(3)

    workbook.save('卡片导入文件.xlsx')

    return True


if __name__ == '__main__':
    if '898602' in ['898600', '898602', '898604', '898607']:
        print('中国移动')
